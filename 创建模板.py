#!/usr/bin/env python3
"""
创建表格模板文件
"""

from openpyxl import Workbook

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "单词表"

# 添加表头
headers = ['英文单词', '英文音标', '词根词缀', '例句', '例句释义', '单词释义']
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 添加示例数据
example_data = [
    ['apple', '/ˈæpl/', 'a-pple', 'I eat an apple every day.', '我每天吃一个苹果。', '苹果'],
    ['banana', '/bəˈnɑːnə/', 'ban-ana', 'Bananas are yellow.', '香蕉是黄色的。', '香蕉'],
    ['cat', '/kæt/', 'cat', 'The cat is sleeping.', '猫在睡觉。', '猫'],
    ['dog', '/dɒɡ/', 'dog', 'Dogs are loyal animals.', '狗是忠诚的动物。', '狗'],
    ['elephant', '/ˈelɪfənt/', 'ele-ph-ant', 'Elephants have long trunks.', '大象有长长的鼻子。', '大象']
]

for row, data in enumerate(example_data, 2):
    for col, value in enumerate(data, 1):
        ws.cell(row=row, column=col, value=value)

# 调整列宽
for col in range(1, len(headers) + 1):
    ws.column_dimensions[chr(64 + col)].width = 20

# 保存模板文件
template_filename = '单词表模板.xlsx'
wb.save(template_filename)
print(f"模板文件创建成功: {template_filename}")
